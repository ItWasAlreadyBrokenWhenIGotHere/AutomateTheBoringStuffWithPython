#!/usr/bin/env python3
"""
Module Docstring
"""

__author__ = "Petri Weckström"
__version__ = "0.1.0"
__license__ = "MIT"

import openpyxl
import os
import logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
# logging.disable(logging.CRITICAL) # This disable all logs as CRITICAL level is highest


def main():
    """ Main entry point of the app """
    logging.debug('Start of the Main entry point...')
    os.chdir('./documents')
    logging.debug('Create new Workbook')
    wb = openpyxl.Workbook()  # creates new Workbook object
    sheet_names = wb.sheetnames  # returns list of sheet names
    logging.debug('Following sheets found from the workbook: ' + str(sheet_names))
    logging.debug('Selecting sheet')
    sheet = wb['Sheet']  # returns WorkSheet object
    logging.debug('Writing to cell A1')
    sheet['A1'] = 42
    logging.debug('Writing to cell A2')
    sheet['A2'] = 'Hello Excel'
    logging.debug('Saving workbook')
    logging.debug('Create new sheet')
    sheet2 = wb.create_sheet()
    sheet_names = wb.sheetnames  # returns list of sheet names
    logging.debug('Following sheets found from the workbook: ' + str(sheet_names))
    logging.debug('Renaming sheet')
    sheet2.title = 'New sheet'
    logging.debug('Adding one more sheet to index 0')
    wb.create_sheet(index=0, title='Fist sheet')
    logging.debug('Save Workbook')
    wb.save('new_example.xlsx')


if __name__ == "__main__":
    """ This is executed when run from the command line """
    main()
