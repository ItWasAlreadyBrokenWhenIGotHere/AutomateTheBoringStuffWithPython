#!/usr/bin/env python3
"""
Module Docstring for example how to read Excel workbook
"""

__author__ = "Petri Weckström"
__version__ = "0.1.0"
__license__ = "MIT"

import openpyxl
import os
import logging
logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
# logging.disable(logging.CRITICAL) # This disable all logs as CRITICAL level is highest


def main():
    """ Main entry point of the app """
    logging.debug('Start of the Main entry point...')
    os.chdir('./documents')

    workbook = openpyxl.load_workbook('example.xlsx')
    sheet_names = workbook.sheetnames  # returns list of sheet names
    # print(sheet_names)
    logging.debug('Following sheets found from the workbook: ' + str(sheet_names))
    sheet = workbook['Sheet1']  # returns WorkSheet object
    cell = str(sheet['A1'].value)  # returns cell object which is converted to string
    print('Value from A1 cell: ' + str(cell))

    cell = str(sheet['B1'].value)
    print('Value from B1 cell: ' + str(cell))

    cell = str(sheet['C1'].value)
    print('Value from C1 cell: ' + cell)

    cell = sheet.cell(row=2, column=2)
    print('Value from B2 cell: ' + str(cell))

    for i in range(1, 8):
        print(i, sheet.cell(row=i, column=2).value)


if __name__ == "__main__":
    """ This is executed when run from the command line """
    main()